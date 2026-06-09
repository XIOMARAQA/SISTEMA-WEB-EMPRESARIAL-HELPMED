from io import BytesIO

from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

EMPRESA = 'Seguricel S.A.C.'
SISTEMA = 'HelpMed'

PROB_LABELS = ['Muy baja', 'Baja', 'Media', 'Alta', 'Muy alta']
IMP_LABELS = ['Insignificante', 'Menor', 'Moderado', 'Mayor', 'Crítico']

NIVEL_TEXTO = {
    'bajo': 'Bajo (1-5)',
    'medio': 'Medio (6-10)',
    'alto': 'Alto (11-15)',
    'critico': 'Crítico (16-25)',
}

COLOR_PRIMARIO = '0D6EFD'
COLOR_TEXTO = '1A2332'

NIVEL_HEX = {
    'bajo': 'BBF7D0',
    'medio': 'FDE68A',
    'alto': 'FDBA74',
    'critico': 'FCA5A5',
}

NIVEL_HEX_SUAVE = {
    'bajo': 'F0FDF4',
    'medio': 'FFFBEB',
    'alto': 'FFF7ED',
    'critico': 'FEF2F2',
}

BORDE_FINO = Side(style='thin', color='DEE2E6')


def valor_a_nivel(valor: int) -> str:
    if valor <= 5:
        return 'bajo'
    if valor <= 10:
        return 'medio'
    if valor <= 15:
        return 'alto'
    return 'critico'


def _nombre_archivo(slug, extension):
    fecha = timezone.now().strftime('%Y%m%d_%H%M')
    return f'helpmed_{slug}_{fecha}.{extension}'


def fecha_generacion():
    return timezone.now().strftime('%d/%m/%Y %H:%M')


# ── PDF ───────────────────────────────────────────────────────────────────────

def respuesta_pdf(elementos, slug: str, horizontal: bool = False) -> HttpResponse:
    buffer = BytesIO()
    pagesize = landscape(A4) if horizontal else A4
    doc = SimpleDocTemplate(
        buffer,
        pagesize=pagesize,
        leftMargin=1.2 * cm,
        rightMargin=1.2 * cm,
        topMargin=1.2 * cm,
        bottomMargin=1.2 * cm,
    )
    doc.build(elementos)
    buffer.seek(0)
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{_nombre_archivo(slug, "pdf")}"'
    return response


def pdf_encabezado(titulo: str, descripcion: str = ''):
    styles = getSampleStyleSheet()
    bloques = [
        Paragraph(titulo, ParagraphStyle(
            'Titulo',
            parent=styles['Heading1'],
            fontSize=16,
            spaceAfter=4,
            textColor=colors.HexColor(f'#{COLOR_TEXTO}'),
            fontName='Helvetica-Bold',
        )),
        Paragraph(
            f'{EMPRESA} — {SISTEMA} · Generado: {fecha_generacion()}',
            ParagraphStyle('Meta', parent=styles['Normal'], fontSize=9, textColor=colors.grey, spaceAfter=6),
        ),
    ]
    if descripcion:
        bloques.append(Paragraph(descripcion, ParagraphStyle(
            'Desc', parent=styles['Normal'], fontSize=8, textColor=colors.HexColor('#64748B'), spaceAfter=10,
        )))
    return bloques


def pdf_seccion(titulo: str):
    styles = getSampleStyleSheet()
    return Paragraph(titulo, ParagraphStyle(
        'Seccion',
        parent=styles['Heading2'],
        fontSize=11,
        spaceBefore=8,
        spaceAfter=6,
        textColor=colors.HexColor(f'#{COLOR_TEXTO}'),
        fontName='Helvetica-Bold',
    ))


def pdf_nota(texto: str):
    styles = getSampleStyleSheet()
    return Paragraph(texto, ParagraphStyle(
        'Nota', parent=styles['Normal'], fontSize=7.5, textColor=colors.HexColor('#64748B'), spaceAfter=8,
    ))


def pdf_tabla(filas: list[list], col_widths=None, resaltar_col=None, niveles_fila=None, **_kwargs):
    """niveles_fila: lista de claves de nivel (bajo/medio/alto/critico) por fila de datos."""
    if not filas:
        filas = [['Sin registros']]
    tabla = Table(filas, colWidths=col_widths, repeatRows=1)
    estilo = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(f'#{COLOR_PRIMARIO}')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#DEE2E6')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8FAFC')]),
    ]
    if resaltar_col is not None and niveles_fila:
        for row_idx, nivel in enumerate(niveles_fila, start=1):
            if nivel and nivel in NIVEL_HEX:
                hex_c = NIVEL_HEX[nivel]
                estilo.append((
                    'BACKGROUND', (resaltar_col, row_idx), (resaltar_col, row_idx),
                    colors.HexColor(f'#{hex_c}'),
                ))
    tabla.setStyle(TableStyle(estilo))
    return tabla


def pdf_leyenda_riesgos():
    items = [
        ('Bajo (1-5)', NIVEL_HEX['bajo']),
        ('Medio (6-10)', NIVEL_HEX['medio']),
        ('Alto (11-15)', NIVEL_HEX['alto']),
        ('Crítico (16-25)', NIVEL_HEX['critico']),
    ]
    filas = [['Leyenda de niveles de riesgo', '', '', '']]
    data = [[nombre, ''] for nombre, _ in items]
    tabla = Table([['']] + data, colWidths=[120, 30])
    # Simpler: use a horizontal legend table
    legend_rows = [['Leyenda:'] + [n for n, _ in items]]
    legend_table = Table(legend_rows, colWidths=[55, 85, 85, 85, 95])
    style_cmds = [
        ('FONTNAME', (0, 0), (0, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
    ]
    for i, (_, hex_c) in enumerate(items, start=1):
        style_cmds.append(('BACKGROUND', (i, 0), (i, 0), colors.HexColor(f'#{hex_c}')))
        style_cmds.append(('BOX', (i, 0), (i, 0), 0.5, colors.HexColor('#CBD5E1')))
    legend_table.setStyle(TableStyle(style_cmds))
    return legend_table


def pdf_mapa_calor(grid: dict):
    """grid: {(impacto, prob): [codigos]}"""
    header = ['Impacto ↓\nProb →', 'P1\nMuy baja', 'P2\nBaja', 'P3\nMedia', 'P4\nAlta', 'P5\nMuy alta']
    filas = [header]
    estilo_cmds = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(f'#{COLOR_PRIMARIO}')),
        ('BACKGROUND', (0, 1), (0, -1), colors.HexColor('#F1F5F9')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CBD5E1')),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]
    for row_idx, imp in enumerate(range(5, 0, -1), start=1):
        fila = [f'I{imp}\n{IMP_LABELS[imp - 1]}']
        for prob in range(1, 6):
            valor = imp * prob
            cods = grid.get((imp, prob), [])
            texto = str(valor)
            if cods:
                texto += '\n' + '\n'.join(cods)
            fila.append(texto)
            nivel = valor_a_nivel(valor)
            tiene = bool(cods)
            hex_c = NIVEL_HEX[nivel] if tiene else NIVEL_HEX_SUAVE[nivel]
            estilo_cmds.append((
                'BACKGROUND', (prob, row_idx), (prob, row_idx), colors.HexColor(f'#{hex_c}'),
            ))
            if tiene:
                estilo_cmds.append(('FONTNAME', (prob, row_idx), (prob, row_idx), 'Helvetica-Bold'))
        filas.append(fila)

    tabla = Table(filas, colWidths=[72, 68, 68, 68, 68, 68])
    tabla.setStyle(TableStyle(estilo_cmds))
    return tabla


# ── Excel ─────────────────────────────────────────────────────────────────────

def respuesta_excel(workbook: Workbook, slug: str) -> HttpResponse:
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{_nombre_archivo(slug, "xlsx")}"'
    return response


def crear_workbook(titulo_hoja: str) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = titulo_hoja[:31]
    return wb


def _fill(hex_color):
    return PatternFill('solid', fgColor=hex_color)


def _font(bold=False, color='000000', size=11):
    return Font(bold=bold, color=color, size=size, name='Calibri')


def excel_portada(ws, titulo: str, secciones: list[str] | None = None, nota: str = ''):
    ws.merge_cells('A1:F1')
    c = ws['A1']
    c.value = titulo
    c.font = _font(bold=True, size=16, color=COLOR_TEXTO)
    ws.merge_cells('A2:F2')
    ws['A2'].value = f'{EMPRESA} — {SISTEMA}'
    ws['A2'].font = _font(size=11, color='64748B')
    ws.merge_cells('A3:F3')
    ws['A3'].value = f'Generado: {fecha_generacion()}'
    ws['A3'].font = _font(size=10, color='64748B')
    fila = 5
    if nota:
        ws.merge_cells(f'A{fila}:F{fila}')
        ws.cell(row=fila, column=1, value=nota).font = _font(size=10, color='475569')
        fila += 2
    if secciones:
        ws.cell(row=fila, column=1, value='Contenido del reporte').font = _font(bold=True, size=12)
        fila += 1
        for s in secciones:
            ws.cell(row=fila, column=1, value=f'• {s}')
            fila += 1
    ws.column_dimensions['A'].width = 18


def excel_encabezado_tabla(ws, fila: int, titulo: str, columnas: list[str]) -> int:
    ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=len(columnas))
    cell = ws.cell(row=fila, column=1, value=titulo)
    cell.font = _font(bold=True, size=12, color=COLOR_TEXTO)
    fila += 2
    for col, nombre in enumerate(columnas, start=1):
        c = ws.cell(row=fila, column=col, value=nombre)
        c.fill = _fill(COLOR_PRIMARIO)
        c.font = _font(bold=True, color='FFFFFF', size=10)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = Border(left=BORDE_FINO, right=BORDE_FINO, top=BORDE_FINO, bottom=BORDE_FINO)
    return fila + 1


def excel_celda_datos(ws, fila, col, valor, nivel=None, align='left'):
    c = ws.cell(row=fila, column=col, value=valor)
    c.alignment = Alignment(horizontal=align, vertical='center', wrap_text=True)
    c.border = Border(left=BORDE_FINO, right=BORDE_FINO, top=BORDE_FINO, bottom=BORDE_FINO)
    if nivel and nivel in NIVEL_HEX:
        c.fill = _fill(NIVEL_HEX[nivel])
        if nivel in ('alto', 'critico'):
            c.font = _font(bold=True, color='FFFFFF' if nivel == 'critico' else '1E293B', size=10)
    return c


def excel_ajustar_columnas(ws, min_width=10, max_width=45):
    for col_cells in ws.columns:
        letter = get_column_letter(col_cells[0].column)
        max_len = 0
        for cell in col_cells:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[letter].width = min(max(max_len + 2, min_width), max_width)


def excel_mapa_calor(ws, fila_inicio: int, grid: dict) -> int:
    ws.merge_cells(start_row=fila_inicio, start_column=1, end_row=fila_inicio, end_column=6)
    ws.cell(row=fila_inicio, column=1, value='Mapa de calor — Impacto × Probabilidad').font = _font(bold=True, size=12)
    fila = fila_inicio + 2

    ws.cell(row=fila, column=1, value='Impacto ↓ / Prob →').font = _font(bold=True, size=9)
    for prob in range(1, 6):
        c = ws.cell(row=fila, column=prob + 1, value=f'P{prob}\n{PROB_LABELS[prob - 1]}')
        c.fill = _fill(COLOR_PRIMARIO)
        c.font = _font(bold=True, color='FFFFFF', size=9)
        c.alignment = Alignment(horizontal='center', wrap_text=True)
    fila += 1

    for imp in range(5, 0, -1):
        lbl = ws.cell(row=fila, column=1, value=f'I{imp}\n{IMP_LABELS[imp - 1]}')
        lbl.font = _font(bold=True, size=9)
        lbl.fill = _fill('F1F5F9')
        lbl.alignment = Alignment(horizontal='center', wrap_text=True)
        for prob in range(1, 6):
            valor = imp * prob
            cods = grid.get((imp, prob), [])
            texto = str(valor)
            if cods:
                texto += '\n' + '\n'.join(cods)
            nivel = valor_a_nivel(valor)
            hex_c = NIVEL_HEX[nivel] if cods else NIVEL_HEX_SUAVE[nivel]
            c = ws.cell(row=fila, column=prob + 1, value=texto)
            c.fill = _fill(hex_c)
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            c.font = _font(bold=bool(cods), size=9)
            c.border = Border(left=BORDE_FINO, right=BORDE_FINO, top=BORDE_FINO, bottom=BORDE_FINO)
        ws.row_dimensions[fila].height = 42
        fila += 1

    fila += 1
    leyenda_fila = fila
    ws.cell(row=leyenda_fila, column=1, value='Leyenda:').font = _font(bold=True, size=9)
    for i, (nombre, hex_c) in enumerate([
        ('Bajo (1-5)', NIVEL_HEX['bajo']),
        ('Medio (6-10)', NIVEL_HEX['medio']),
        ('Alto (11-15)', NIVEL_HEX['alto']),
        ('Crítico (16-25)', NIVEL_HEX['critico']),
    ], start=2):
        c = ws.cell(row=leyenda_fila, column=i, value=nombre)
        c.fill = _fill(hex_c)
        c.alignment = Alignment(horizontal='center')
        c.font = _font(size=9)
    return fila + 2
